from io import BytesIO
from openpyxl.utils import get_column_letter
import openpyxl
from flask import Flask, request, jsonify, session, send_file, render_template
from flask_cors import CORS
from controllers.AdminManager import AdminManager
from controllers.TestManager import TestManager
from controllers.UserManager import UserManager

app = Flask(__name__, template_folder='templates', static_folder='static')
allowed_origins = ["*"]

app.config.update(SESSION_COOKIE_SECURE=True, SESSION_COOKIE_HTTPONLY=True, SESSION_COOKIE_SAMESITE='None',
                  PERMANENT_SESSION_LIFETIME=86400)

CORS(app, supports_credentials=True, origins=allowed_origins)
app.secret_key = 'URFU-INNOVATE-2024'
deploy = False
if deploy:
    db_path = '/home/urfuinnovate/URFU-Innovation/backend/data/innovate.db3'
    belbin_test = '/home/urfuinnovate/URFU-Innovation/backend/data/belbin/belbin.json'
    admin_manager = AdminManager(db_path)
    user_manager = UserManager(db_path)
    test_manager = TestManager(db_path, belbin_test)
else:
    admin_manager = AdminManager()
    user_manager = UserManager()
    test_manager = TestManager()


@app.route('/api/register-user', methods=['POST'])
def register_user():
    data = request.form
    full_name = data.get('full_name')
    phone_number = data.get('phone_number')
    telegram_id = data.get('telegram_id')
    if not all([full_name, phone_number, telegram_id]):
        return jsonify({"success": False, "message": "Необходимо заполнить все поля"}), 400

    is_success, message = user_manager.register_user(full_name, phone_number, telegram_id)
    session['telegram_id'] = telegram_id

    return jsonify({"success": is_success, "message": 'Форма успешно принята'}), 201


@app.route('/api/logout-user', methods=['POST'])
def logout_user():
    session.pop('telegram_id', None)
    return jsonify({"success": True, "message": 'Вы успешно вышли'}), 200


@app.route('/api/get-all-test', methods=['GET'])
def get_all_test():
    tests = test_manager.get_tests()
    corr_tests = []
    for test in tests:
        corr_test = {'test_title': test[0], 'test_url': test[1]}
        corr_tests.append(corr_test)
    return jsonify(corr_tests)


def get_top_sections(section_data):
    sorted_sections = sorted(section_data.items(), key=lambda x: x[1], reverse=True)
    top_values = [sorted_sections[0][1], sorted_sections[1][1], sorted_sections[2][1]]

    if len(set(top_values)) == 3:
        return [sorted_sections[0][0], sorted_sections[1][0]]
    else:
        return [sorted_sections[0][0], sorted_sections[1][0], sorted_sections[2][0]]


@app.route('/api/belbin-test', methods=['GET', 'POST'])
def processing_form():
    if request.method == 'POST':
        return process_post_request()

    if request.method == 'GET':
        return get_questions()


def process_post_request():
    data = request.json
    telegram_id = session.get('telegram_id')
    if not telegram_id:
        return jsonify({"success": False, "message": "Пользователь не авторизован!"}), 401

    user_id = user_manager.get_user_id(telegram_id)[0]
    test_id = 1
    user_test_id = user_manager.add_test_to_user(user_id, test_id)

    result = calculate_section_scores(data)

    roles_data = test_manager.get_roles_and_descriptions()
    data_percentages = calculate_percentages(result)
    data_percentages = dict(sorted(data_percentages.items(), key=lambda item: item[1], reverse=True))
    top_result, bottom_result = get_tops_and_bottoms_sections(data_percentages)
    built_top_result = build_top_result(top_result, roles_data)
    built_bottom_result = build_bottom_result(bottom_result, roles_data)
    test_manager.save_user_answers(user_test_id, data_percentages)
    data_roles = {roles_data.get(k).get('role_in_team'): v for k, v in data_percentages.items()}

    logout_user()

    return jsonify({
        "success": True,
        "message": "Форма успешно принята",
        "prefer_roles": built_top_result,
        "un_prefer_roles": built_bottom_result,
        "all_roles": data_roles
    }), 200


def get_questions():
    questions = test_manager.get_all_questions()
    return jsonify({"success": True, "questions": questions}), 200


def calculate_section_scores(data):
    return {
        'section1': data[2][0] + data[1][1] + data[5][2] + data[0][3] + data[4][5] + data[6][6] + data[3][7],
        'section2': data[6][0] + data[3][1] + data[2][2] + data[4][3] + data[1][4] + data[0][5] + data[5][6],
        'section3': data[5][0] + data[0][2] + data[2][3] + data[3][4] + data[6][5] + data[1][6] + data[4][7],
        'section4': data[4][0] + data[6][1] + data[3][2] + data[1][3] + data[5][4] + data[2][6] + data[0][7],
        'section5': data[1][0] + data[4][1] + data[3][3] + data[6][4] + data[5][5] + data[0][6] + data[2][7],
        'section6': data[3][0] + data[0][1] + data[5][1] + data[4][2] + data[2][4] + data[1][5] + data[6][7],
        'section7': data[0][0] + data[1][2] + data[6][3] + data[4][4] + data[2][5] + data[3][6] + data[5][7],
        'section8': data[2][1] + data[6][2] + data[5][3] + data[0][4] + data[3][5] + data[4][6] + data[1][7],
    }


def calculate_percentages(result):
    total_sum = sum(result.values())
    return {key: round((value / total_sum) * 100) for key, value in result.items()}


def get_tops_and_bottoms_sections(data_percentages):
    filtered_data = {key: value for key, value in data_percentages.items() if value > 0}
    sorted_values = sorted(set(filtered_data.values()), reverse=True)
    top_two_values = set(sorted_values[:2])
    top_result = {key: value for key, value in filtered_data.items() if value in top_two_values}

    if len(top_result) < 3:
        top_two_values = set(sorted_values[:3])
        top_result = {key: value for key, value in filtered_data.items() if value in top_two_values}

    filtered_data = {key: value for key, value in data_percentages.items() if key not in top_result.keys()}
    sorted_values_bottom = sorted(set(filtered_data.values()))
    bottom_two_values = set(sorted_values_bottom[:2])
    bottom_result = {key: value for key, value in filtered_data.items() if
                     value in bottom_two_values and key not in top_result}

    if len(bottom_result) < 3:
        bottom_two_values = set(sorted_values_bottom[:3])
        bottom_result = {key: value for key, value in filtered_data.items() if
                         value in bottom_two_values and key not in top_result}

    return top_result, bottom_result


def build_top_result(final_data, roles_data):
    final_result = []
    for section, value in final_data.items():
        role_info = roles_data.get(section)
        if role_info:
            final_result.append({
                'role': role_info['role_in_team'],
                'strong_side': role_info['strong-side'],
                'value': value,
                'description': role_info['description'],
                'file_name': role_info['file_name']
            })
    return final_result


def build_bottom_result(final_data, roles_data):
    final_result = []
    for section, value in final_data.items():
        role_info = roles_data.get(section)
        if role_info:
            final_result.append({
                'role': role_info['role_in_team'],
                'value': value,
                'weak_side': role_info['weak-side'],
                'recommendations': role_info['recommendations'],
                'description': role_info['description'],
                'file_name': role_info['file_name']
            })
    return final_result


@app.route('/api/get-test-results', methods=['GET'])
def get_test_results():
    data = request.form
    telegram_id = data.get('telegram_id')
    test_name = data.get('test_name')
    start_date = data.get('start_date')
    end_date = data.get('end_date')

    current_user = session.get('admin_username')
    if not current_user or not admin_manager.is_admin(current_user):
        return jsonify({"success": False, "message": "Только администратторы могут выполнять данное действие!"}), 403

    results = test_manager.get_filtered_results(telegram_id, test_name, start_date, end_date)
    return jsonify({"success": True, "results": results}), 200


@app.route('/api/save-test-results', methods=['GET'])
def save_test_results():
    data = request.form
    telegram_id = data.get('telegram_id')
    test_name = data.get('test_name')
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    current_user = session.get('admin_username')
    if not current_user or not admin_manager.is_admin(current_user):
        return jsonify({"success": False, "message": "Только администраторы могут выполнять данное действие!"}), 403

    results = test_manager.get_filtered_results(telegram_id, test_name, start_date, end_date)
    wb, ws = create_excel_file(results)
    adjust_column_widths(ws)
    return save_and_send_file(wb)


def create_excel_file(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Test Results'
    keys = list(results[0].get('sections').keys())
    headers = ['ФИО', 'Номер телефона', 'Telegram ID', 'Название теста', 'Время прохождения'] + keys
    ws.append(headers)

    for result in results:
        row = format_result_row(result)
        ws.append(row)

    return wb, ws


def format_result_row(result):
    full_name = result.get('full_name')
    phone_number = result.get('phone_number')
    telegram_id = result.get('telegram_id')
    test_name = result.get('test_name')
    timestamp = result.get('timestamp')
    sections = result.get('sections')
    sections_str = [float(i) for i in sections.values()]
    return [full_name, phone_number, telegram_id, test_name, timestamp] + sections_str


def adjust_column_widths(ws):
    for col in range(1, len(ws[1]) + 1):
        column = get_column_letter(col)
        max_length = max(len(str(cell.value)) for cell in ws[column] if cell.value is not None)
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width


def save_and_send_file(wb):
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return send_file(file_stream, as_attachment=True, download_name='test_results.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/register', methods=['POST'])
def register():
    data = request.form
    current_user = session.get('admin_username')
    if not current_user:
        return jsonify({"success": False, "message": "Пользователь не авторизован!"}), 401

    username = data.get('username')
    password1 = data.get('password1')
    password2 = data.get('password2')

    is_success, message = admin_manager.register_admin(current_user, username, password1, password2)
    code = 201 if is_success else 403
    return jsonify({"success": is_success, "message": message}), code


@app.route('/api/login', methods=['POST'])
def login():
    data = request.form
    username = data.get('username')
    password = data.get('password')
    is_success, message = admin_manager.login_admin(username, password)
    if not is_success:
        return jsonify({"success": False, "message": message}), 401
    session['admin_username'] = username
    role = admin_manager.is_super_admin(username)
    return jsonify({"success": is_success, "message": message, "super_admin": role}), 200


@app.route('/api/change-password', methods=['POST'])
def change_password():
    data = request.form
    username = session.get('username')
    current_password = data.get('current_password')
    new_password1 = data.get('new_password1')
    new_password2 = data.get('new_password2')

    if not username:
        return jsonify({"success": False, "message": "Пользователь не авторизован!"}), 401

    is_success, message = admin_manager.change_password(username, current_password, new_password1, new_password2)
    code = 200 if is_success else 400
    return jsonify({"success": is_success, "message": message}), code


@app.route('/api/logout', methods=['GET'])
def logout():
    session.pop('admin_username', None)
    return jsonify({"success": True, "message": 'Вы успешно вышли'}), 200


@app.route('/api/delete_admin', methods=['DELETE'])
def delete_admin():
    data = request.form
    username = data.get('username')
    current_user = session.get('admin_username')
    is_success, message = admin_manager.delete_admin(current_user, username)
    code = 200 if is_success else 403
    return jsonify({"success": is_success, "message": message}), code


@app.route('/api/admins', methods=['GET'])
def get_all_admins():
    current_user = session.get('admin_username')

    is_success, result = admin_manager.get_all_admins(current_user)
    if not is_success:
        return jsonify({"success": False, "message": result}), 403 if "Только супер-администраторы" in result else 401

    return jsonify({"success": True, "admins": result}), 200


@app.route('/api/promote-to-super-admin', methods=['POST'])
def promote_to_super_admin():
    current_user = session.get('admin_username')
    if not current_user:
        return jsonify({"success": False, "message": "Пользователь не авторизован!"}), 401

    data = request.form
    username = data.get('username')

    is_success, message = admin_manager.promote_to_super_admin(current_user, username)
    code = 200 if is_success else 403
    return jsonify({"success": is_success, "message": message}), code


@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def index(path):
    return render_template('index.html')


if __name__ == '__main__':
    # app.run(host='localhost', port=5000, debug=False)
    app.run()
