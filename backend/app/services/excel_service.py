from io import BytesIO

import openpyxl
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def create_excel_file(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Test Results'
    print(results)
    keys = list(results[0].get('sections').keys())
    headers = ['ФИО', 'Номер телефона', 'Telegram ID', 'Название теста',
               'Время прохождения'] + keys
    ws.append(headers)

    for result in results:
        row = format_result_row(result)
        ws.append(row)

    format_phone_numbers(ws)
    highlight_cells(ws, len(headers))  # Применяем стили
    return wb, ws


def format_result_row(result):
    full_name = result.get('full_name')
    phone_number = str(result.get('phone_number'))  # Преобразуем в строку
    telegram_id = result.get('telegram_id')
    test_name = result.get('test_name')
    timestamp = result.get('timestamp')
    sections = result.get('sections')
    sections_values = [float(i) for i in sections.values()]
    return [full_name, phone_number, telegram_id, test_name,
            timestamp] + sections_values


def format_phone_numbers(ws):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2,
                            max_col=2):
        for cell in row:
            cell.number_format = '@'  # Устанавливаем текстовый формат


def highlight_cells(ws, num_columns):
    green_fill = PatternFill(start_color="FF99FFCC", end_color="FF99FFCC",
                             fill_type="solid")
    red_fill = PatternFill(start_color="FFFF7C80", end_color="FFFF7C80",
                           fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6,
                            max_col=num_columns):
        scores = [cell.value for cell in row if
                  isinstance(cell.value, (int, float))]
        if not scores:
            continue

        top_3_values = sorted(scores, reverse=True)[:3]

        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value in top_3_values:
                    cell.fill = green_fill
                elif 0 <= cell.value <= 3:
                    cell.fill = red_fill


def adjust_column_widths(ws):
    for col in range(1, len(ws[1]) + 1):
        column = get_column_letter(col)
        max_length = max(len(str(cell.value)) for cell in ws[column] if cell.value is not None)
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width


def save_and_send_file(wb):
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return send_file(file_stream, as_attachment=True, download_name='test_results.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def create_disc_excel_file(data: list):
    """
    data: список словарей с ключами:
        user_test_id, full_name, phone_number, telegram_id, test_name, timestamp, D, I, S, C
    возвращает: (workbook, worksheet)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты DISC"

    # Заголовки на русском
    headers = [
        "ФИО",
        "Номер телефона",
        "Telegram ID",
        "Название теста",
        "Дата и время",
        "D",
        "I",
        "S",
        "C"
    ]
    ws.append(headers)

    # Данные
    for row in data:
        ws.append([
            row.get("full_name", ""),
            row.get("phone_number", ""),
            row.get("telegram_id", ""),
            row.get("test_name", ""),
            row.get("timestamp", ""),
            row.get("D", ""),
            row.get("I", ""),
            row.get("S", ""),
            row.get("C", "")
        ])

    return wb, ws