from datetime import datetime, timedelta

from flask import Blueprint, request, session, jsonify
from flask import session as flask_session
from openpyxl.styles import Border, Side
from sqlalchemy.orm import joinedload

from ..controllers.AdminManager import AdminManager
from ..controllers.TestManager import TestManager, get_tests
from ..controllers.UserManager import UserManager
from ..database import SessionLocal
from ..models.test import TestEnum
from ..models.user import User
from ..models.user_test import UserTest
from ..models.user_test_result import UserTestResult
from ..paths import BELBIN_TEST_PATH
from ..services.belbin_service import process_post_request, get_questions
from ..services.disc_service import get_disc_test_questions, calculate_disc_scores, save_disc_test_results
from ..services.excel_service import create_excel_file, adjust_column_widths, save_and_send_file, \
    create_disc_excel_file, highlight_belbin_cells, auto_adjust_column_width

test_bp = Blueprint('test', __name__)
user_manager = UserManager(session_factory=SessionLocal)
test_manager = TestManager(session_factory=SessionLocal, belbin_test=BELBIN_TEST_PATH)


@test_bp.route('/get-all-test', methods=['GET'])
def get_all_test():
    tests = get_tests()
    return jsonify([{'test_title': t[0], 'test_url': t[1]} for t in tests])


@test_bp.route('/belbin-test', methods=['GET', 'POST'])
def processing_form():
    if request.method == 'POST':
        return process_post_request(user_manager, test_manager, flask_session)
    return get_questions(test_manager)


@test_bp.route('/disc-test', methods=['GET', 'POST'])
def disc_test():
    if request.method == 'POST':
        data = request.get_json()
        telegram_id = session.get('telegram_id')
        if not telegram_id:
            return jsonify(
                {"success": False, "message": "Пользователь не авторизован!"}), 401
        if not data or "answers" not in data:
            return {"error": "Invalid payload"}, 400

        result = calculate_disc_scores(data)
        user_id = user_manager.get_user_id(telegram_id)
        save_disc_test_results(
            db=SessionLocal(),
            user_id=user_id,
            test_id=TestEnum.DISC,
            scores=result["scores"]
        )
        return result, 200

    return get_disc_test_questions()


@test_bp.route('/get-test-results', methods=['GET'])
def get_test_results():
    telegram_id = request.args.get('telegram_id')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    test_name = "BELBIN"
    current_user = session.get('admin_username')
    from ..controllers.AdminManager import AdminManager
    admin_manager = AdminManager(SessionLocal)
    if not current_user or not admin_manager.is_admin(current_user):
        return jsonify({"success": False, "message": "Только администраторы могут выполнять данное действие!"}), 403
    results = test_manager.get_filtered_results(telegram_id, test_name, start_date, end_date)
    return jsonify({"success": True, "results": results}), 200


@test_bp.route('/get-test-results', methods=['DELETE'])
def admin_delete_belbin():
    current_user = flask_session.get('admin_username')
    admin_manager = AdminManager(SessionLocal)
    if not current_user or not admin_manager.is_admin(current_user):
        return jsonify({"success": False, "message": "Только администраторы могут выполнять данное действие!"}), 403

    start_date_str = request.args.get('start_date')  # YYYY-MM-DD
    end_date_str = request.args.get('end_date')  # YYYY-MM-DD

    if not start_date_str and not end_date_str:
        return jsonify({"success": False, "message": "Не указана ни start_date, ни end_date"}), 400

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d") if start_date_str else None
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d") if end_date_str else None
    except ValueError:
        return jsonify({"success": False, "message": "Неверный формат даты"}), 400

    with SessionLocal() as db_session:
        query = db_session.query(UserTest)

        if start_date and not end_date:
            query = query.filter(UserTest.timestamp >= start_date,
                                 UserTest.test_id == TestEnum.BELBIN)

        elif end_date and not start_date:
            end_date = end_date + timedelta(days=1)
            query = query.filter(UserTest.timestamp <= end_date,
                                 UserTest.test_id == TestEnum.BELBIN)

        elif start_date and end_date:
            end_date = end_date + timedelta(days=1)
            query = query.filter(UserTest.timestamp >= start_date, UserTest.timestamp <= end_date,
                                 UserTest.test_id == TestEnum.BELBIN)

        tests_to_delete = query.all()
        deleted_count = 0

        for test in tests_to_delete:
            # Сначала удаляем результаты
            db_session.query(UserTestResult).filter_by(user_test_id=test.user_test_id).delete()
            # Потом сам тест
            db_session.delete(test)
            deleted_count += 1

        db_session.commit()

        return jsonify({"success": True, "deleted_tests": deleted_count}), 200


@test_bp.route('/disc-test-results', methods=['GET'])
def admin_disc_test_results():
    current_user = flask_session.get('admin_username')
    admin_manager = AdminManager(SessionLocal)
    if not current_user or not admin_manager.is_admin(current_user):
        return jsonify({"success": False, "message": "Только администраторы могут выполнять данное действие!"}), 403

    telegram_id = request.args.get('telegram_id')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    try:
        start_date = datetime.strptime(start_date, "%Y-%m-%d") if start_date else None
        end_date = datetime.strptime(end_date, "%Y-%m-%d") if end_date else None
    except ValueError:
        return jsonify({"success": False, "message": "Неверный формат даты"}), 400

    with SessionLocal() as db_session:
        query = (
            db_session
            .query(UserTest, User, UserTestResult)
            .join(User, User.user_id == UserTest.user_id)
            .join(UserTestResult, UserTestResult.user_test_id == UserTest.user_test_id)
        )

        if telegram_id:
            query = query.filter(User.telegram_id == telegram_id)

        if start_date:
            query = query.filter(UserTest.timestamp >= start_date)

        if end_date:
            end_date += timedelta(days=1)
            query = query.filter(UserTest.timestamp < end_date)

        table = {}
        for user_test, user, result in query.all():
            key = user_test.user_test_id
            if key not in table:
                table[key] = {
                    "test_name": user_test.test_id.display_name,
                    "telegram_id": user.telegram_id,
                    "full_name": user.full_name,
                    "phone_number": user.phone_number,
                    "timestamp": user_test.timestamp.strftime(
                        "%a, %d %b %Y %H:%M:%S GMT"
                    ),
                    "sections": {}
                }
            table[key]["sections"][result.scale] = result.value

        table_list = list(table.values())

        return jsonify({"success": True, "results": table_list}), 200


@test_bp.route('/disc-test-results', methods=['DELETE'])
def admin_delete_disc_tests():
    current_user = flask_session.get('admin_username')
    admin_manager = AdminManager(SessionLocal)
    if not current_user or not admin_manager.is_admin(current_user):
        return jsonify({"success": False, "message": "Только администраторы могут выполнять данное действие!"}), 403

    start_date_str = request.args.get('start_date')  # YYYY-MM-DD
    end_date_str = request.args.get('end_date')  # YYYY-MM-DD

    if not start_date_str and not end_date_str:
        return jsonify({"success": False, "message": "Не указана ни start_date, ни end_date"}), 400

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d") if start_date_str else None
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d") if end_date_str else None
    except ValueError:
        return jsonify({"success": False, "message": "Неверный формат даты"}), 400

    with SessionLocal() as db_session:
        query = db_session.query(UserTest)

        if start_date and not end_date:
            query = query.filter(UserTest.timestamp >= start_date,
                                 UserTest.test_id == TestEnum.DISC)
        elif end_date and not start_date:
            end_date = end_date + timedelta(days=1)
            query = query.filter(UserTest.timestamp < end_date,
                                 UserTest.test_id == TestEnum.DISC)

        elif start_date and end_date:
            end_date = end_date + timedelta(days=1)
            query = query.filter(UserTest.timestamp >= start_date, UserTest.timestamp <= end_date,
                                 UserTest.test_id == TestEnum.DISC)

        tests_to_delete = query.all()
        deleted_count = 0

        for test in tests_to_delete:
            # Сначала удаляем результаты
            db_session.query(UserTestResult).filter_by(user_test_id=test.user_test_id).delete()
            # Потом сам тест
            db_session.delete(test)
            deleted_count += 1

        db_session.commit()

        return jsonify({"success": True, "deleted_tests": deleted_count}), 200


@test_bp.route('/user-tests', methods=['GET'])
def get_user_tests():
    telegram_id = request.args.get('telegram_id')
    test_name = request.args.get('test_name')  # новый параметр
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    current_user = session.get('admin_username')
    from ..controllers.AdminManager import AdminManager
    admin_manager = AdminManager(SessionLocal)
    if not current_user or not admin_manager.is_admin(current_user):
        return jsonify({"success": False, "message": "Только администраторы могут выполнять данное действие!"}), 403

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d") if start_date_str else None
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d") + timedelta(days=1) if end_date_str else None
    except ValueError:
        return jsonify({"success": False, "message": "Неверный формат даты"}), 400

    users_dict = {}

    from ..controllers.TestManager import TestManager
    test_manager = TestManager(SessionLocal)
    roles_dict = test_manager.get_roles_and_descriptions()
    belbin_roles = [v["role_in_team"] for v in roles_dict.values()]

    with SessionLocal() as db_session:
        query = db_session.query(UserTest).options(
            joinedload(UserTest.user),
            joinedload(UserTest.results),  # DISC
            joinedload(UserTest.answers)  # BELBIN
        ).join(User)

        if telegram_id:
            query = query.filter(User.telegram_id == telegram_id)

        if start_date:
            query = query.filter(UserTest.timestamp >= start_date)
        if end_date:
            query = query.filter(UserTest.timestamp < end_date)

        if test_name:
            try:
                enum_value = TestEnum[test_name]
                query = query.filter(UserTest.test_id == enum_value)
            except KeyError:
                return jsonify({"success": False, "message": f"Тест {test_name} не найден"}), 400

        query = query.order_by(UserTest.timestamp.desc())
        tests = query.all()

        for ut in tests:
            user = ut.user
            key = user.telegram_id
            if key not in users_dict:
                users_dict[key] = {
                    "telegram_id": user.telegram_id,
                    "full_name": user.full_name,
                    "phone_number": user.phone_number,
                    "tests": []
                }

            test_data = {
                "test_name": getattr(ut.test_id, "display_name", str(ut.test_id)),
                "timestamp": ut.timestamp.strftime("%a, %d %b %Y %H:%M:%S GMT"),
                "sections": {}
            }

            if ut.test_id.name == "DISC" and hasattr(ut, "results") and ut.results:
                test_data["sections"] = {r.scale: r.value for r in ut.results}

            elif ut.test_id.name == "BELBIN" and hasattr(ut, "answers") and ut.answers:
                test_data["sections"] = {
                    role: getattr(ut.answers, f"section{i + 1}")
                    for i, role in enumerate(belbin_roles)
                }

            users_dict[key]["tests"].append(test_data)

    return jsonify({"success": True, "results": list(users_dict.values())}), 200


@test_bp.route('/user-tests', methods=['DELETE'])
def delete_user_tests():
    telegram_id = request.args.get('telegram_id')
    test_name = request.args.get('test_name')  # новый параметр
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    current_user = session.get('admin_username')
    from ..controllers.AdminManager import AdminManager
    admin_manager = AdminManager(SessionLocal)
    if not current_user or not admin_manager.is_admin(current_user):
        return jsonify({"success": False, "message": "Только администраторы могут выполнять данное действие!"}), 403

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d") if start_date_str else None
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d") + timedelta(days=1) if end_date_str else None
    except ValueError:
        return jsonify({"success": False, "message": "Неверный формат даты"}), 400

    with SessionLocal() as db_session:
        query = db_session.query(UserTest).join(User)

        if telegram_id:
            query = query.filter(User.telegram_id == telegram_id)
        if start_date:
            query = query.filter(UserTest.timestamp >= start_date)
        if end_date:
            query = query.filter(UserTest.timestamp < end_date)
        if test_name:
            from ..models.test import TestEnum
            try:
                enum_value = TestEnum[test_name]
                query = query.filter(UserTest.test_id == enum_value)
            except KeyError:
                return jsonify({"success": False, "message": f"Тест {test_name} не найден"}), 400

        tests_to_delete = query.all()
        deleted_count = 0

        for test in tests_to_delete:
            db_session.delete(test)
            deleted_count += 1

        db_session.commit()

    return jsonify({"success": True, "deleted_tests": deleted_count}), 200


@test_bp.route('/save-test-results', methods=['GET'])
def save_test_results():
    data = request.form
    telegram_id = data.get('telegram_id')
    start_date = data.get('start_date')
    end_date = data.get('end_date')

    test_name = "BELBIN"
    current_user = session.get('admin_username')
    from ..controllers.AdminManager import AdminManager
    admin_manager = AdminManager(SessionLocal)
    if not current_user or not admin_manager.is_admin(current_user):
        return jsonify({"success": False, "message": "Только администраторы могут выполнять данное действие!"}), 403

    results = test_manager.get_filtered_results(telegram_id, test_name, start_date, end_date)
    wb, ws = create_excel_file(results)
    adjust_column_widths(ws)
    return save_and_send_file(wb)


@test_bp.route('/save-disc-test-results', methods=['GET'])
def save_disc_test_results_excel():
    current_user = flask_session.get('admin_username')
    admin_manager = AdminManager(SessionLocal)
    if not current_user or not admin_manager.is_admin(current_user):
        return jsonify({"success": False, "message": "Только администраторы могут выполнять данное действие!"}), 403

    telegram_id = request.args.get('telegram_id')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d") if start_date_str else None
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d") if end_date_str else None
    except ValueError:
        return jsonify({"success": False, "message": "Неверный формат даты"}), 400

    with SessionLocal() as db_session:
        query = db_session.query(UserTest, User, UserTestResult).join(User).join(UserTestResult)

        if telegram_id:
            query = query.filter(User.telegram_id == telegram_id)
        if start_date:
            query = query.filter(UserTest.timestamp >= start_date)
        if end_date:
            end_date += timedelta(days=1)
            query = query.filter(UserTest.timestamp < end_date)

        table = {}
        for user_test, user, result in query.all():
            key = user_test.user_test_id
            if key not in table:
                table[key] = {
                    "full_name": user.full_name,
                    "phone_number": user.phone_number,
                    "telegram_id": user.telegram_id,
                    "test_name": user_test.test_id.display_name,
                    "timestamp": user_test.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                    "D": "",
                    "I": "",
                    "S": "",
                    "C": ""
                }
            table[key][result.scale] = result.value

        excel_data = list(table.values())

        wb, ws = create_disc_excel_file(excel_data)
        adjust_column_widths(ws)
        return save_and_send_file(wb)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

@test_bp.route('/user-tests/export', methods=['GET'])
def export_user_tests_excel():
    current_user = session.get('admin_username')

    from ..controllers.TestManager import TestManager
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from flask import send_file
    import tempfile

    # ---- CHECK ADMIN (если нужно) ----
    # admin_manager = AdminManager(SessionLocal)
    # if not current_user or not admin_manager.is_admin(current_user):
    #     return jsonify({"success": False, "message": "Только администраторы"}), 403

    test_manager = TestManager(SessionLocal)
    roles_dict = test_manager.get_roles_and_descriptions()
    belbin_roles = [v["role_in_team"] for v in roles_dict.values()]
    disc_roles = [
        "D\n(доминирование)",
        "I\n(влияние)",
        "S\n(устойчивость)",
        "C\n(сознательность)"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "User tests"

    # ---------- HEADER (2 ROWS) ----------
    ws.append([
        "ФИО", "Номер телефона", "Telegram ID",
        *["BELBIN"] * len(belbin_roles),
        *["DISC"] * len(disc_roles)
    ])

    disc_scales = ["D", "I", "S", "C"]

    ws.append([
        "", "", "",
        *belbin_roles,
        *disc_roles
    ])

    # Объединяем пользовательские поля
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)

    # Объединяем BELBIN
    belbin_start = 4
    belbin_end = belbin_start + len(belbin_roles) - 1
    ws.merge_cells(start_row=1, start_column=belbin_start, end_row=1, end_column=belbin_end)

    # Объединяем DISC
    disc_start = belbin_end + 1
    disc_end = disc_start + len(disc_roles) - 1
    ws.merge_cells(start_row=1, start_column=disc_start, end_row=1, end_column=disc_end)

    # Центрирование
    for col in range(1, disc_end + 1):
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # ---------- DATA ----------
    users_map = {}

    with SessionLocal() as db:
        tests = db.query(UserTest).options(
            joinedload(UserTest.user),
            joinedload(UserTest.results),
            joinedload(UserTest.answers)
        ).all()

        for ut in tests:
            user = ut.user
            uid = user.user_id

            if uid not in users_map:
                users_map[uid] = {
                    "full_name": user.full_name,
                    "phone_number": user.phone_number,
                    "telegram_id": user.telegram_id,
                    "belbin": {},
                    "disc": {}
                }

            if ut.test_id.name == "BELBIN" and ut.answers:
                users_map[uid]["belbin"] = {
                    role: getattr(ut.answers, f"section{i + 1}")
                    for i, role in enumerate(belbin_roles)
                }

            elif ut.test_id.name == "DISC" and ut.results:
                users_map[uid]["disc"] = {
                    r.scale: r.value for r in ut.results
                }

    # ---------- ROWS ----------
    for user in users_map.values():
        row = [
            user["full_name"],
            user["phone_number"],
            user["telegram_id"],
        ]

        # BELBIN
        for role in belbin_roles:
            row.append(user["belbin"].get(role))


        # DISC
        for scale in disc_scales:
            row.append(user["disc"].get(scale))

        ws.append(row)

    belbin_start_col = 4
    belbin_end_col = belbin_start_col + len(belbin_roles) - 1

    highlight_belbin_cells(ws, belbin_start_col, belbin_end_col)
    # ---------- SAVE ----------
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    auto_adjust_column_width(ws)
    wb.save(tmp.name)

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name="user_tests.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
